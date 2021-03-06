"""
    pyexcel_xlsx
    ~~~~~~~~~~~~~~~~~~~

    The lower level xlsx file format handler using openpyxl

    :copyright: (c) 2015-2016 by Onni Software Ltd & its contributors
    :license: New BSD License
"""
import sys
import openpyxl

from pyexcel_io.book import BookReader, BookWriter
from pyexcel_io.sheet import SheetReader, SheetWriter

PY2 = sys.version_info[0] == 2
if PY2 and sys.version_info[1] < 7:
    from ordereddict import OrderedDict
else:
    from collections import OrderedDict    


COLUMNS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
COLUMN_LENGTH = 26


def get_columns(index):
    if index < COLUMN_LENGTH:
        return COLUMNS[index]
    else:
        return (get_columns(int(index // COLUMN_LENGTH) - 1) + COLUMNS[index % COLUMN_LENGTH])


class XLSXSheet(SheetReader):
    """
    xls sheet
    """
    @property
    def name(self):
        return self.native_sheet.title

    def number_of_rows(self):
        """
        Number of rows in the xls sheet
        """
        return self.native_sheet.max_row

    def number_of_columns(self):
        """
        Number of columns in the xls sheet
        """
        return self.native_sheet.max_column

    def cell_value(self, row, column):
        """
        Random access to the xls cells
        """
        actual_row = row + 1
        cell_location = "%s%d" % (get_columns(column), actual_row)
        return self.native_sheet.cell(cell_location).value

    def to_array(self):
        for r in range(0, self.number_of_rows()):
            row = []
            tmp_row = []
            for c in range(0, self.number_of_columns()):
                cell_value = self.cell_value(r, c)
                tmp_row.append(cell_value)
                if cell_value is not None and cell_value != '':
                    row += tmp_row
                    tmp_row = []
            yield row


class XLSXBook(BookReader):
    """
    XLSBook reader

    It reads xls, xlsm, xlsx work book
    """
    def __init__(self):
        BookReader.__init__(self)
        self.book = None

    def open(self, file_name, **keywords):
        BookReader.open(self, file_name, **keywords)
        self._load_from_file()

    def open_stream(self, file_stream, **keywords):
        BookReader.open_stream(self, file_stream, **keywords)
        self._load_from_memory()

    def read_sheet_by_name(self, sheet_name):
        sheet = self.book.get_sheet_by_name(sheet_name)
        if sheet is None:
            raise ValueError("%s cannot be found" % sheet_name)
        else:
            sheet = XLSXSheet(sheet)
            return {sheet_name: sheet.to_array()}

    def read_sheet_by_index(self, sheet_index):
        names = self.book.sheetnames
        length = len(names)
        if sheet_index < length:
            return self.read_sheet_by_name(names[sheet_index])
        else:
            raise IndexError("Index %d of out bound %d" %(
                sheet_index,
                length))

    def read_all(self):
        result = OrderedDict()
        for sheet in self.book:
            sheet = XLSXSheet(sheet)
            result[sheet.name] = sheet.to_array()
        return result
        
    def _load_from_memory(self):
        self.book =  openpyxl.load_workbook(filename=self.file_stream,
                                            data_only=True)

    def _load_from_file(self):
        self.book = openpyxl.load_workbook(filename=self.file_name,
                                           data_only=True)


class XLSXSheetWriter(SheetWriter):
    """
    xls, xlsx and xlsm sheet writer
    """
    def set_sheet_name(self, name):
        self.native_sheet.title = name
        self.current_row = 1

    def write_row(self, array):
        """
        write a row into the file
        """
        for i in range(0, len(array)):
            cell_location = "%s%d" % (get_columns(i), self.current_row)
            self.native_sheet.cell(cell_location).value = array[i]
        self.current_row += 1


class XLSXWriter(BookWriter):
    """
    xls, xlsx and xlsm writer
    """
    def __init__(self):
        BookWriter.__init__(self)
        self.current_sheet = 0
        self.native_book = None

    def open(self, file_name, **keywords):
        BookWriter.open(self, file_name, **keywords)
        self.native_book = openpyxl.Workbook()

    def create_sheet(self, name):
        if self.current_sheet == 0:
            self.current_sheet = 1
            return XLSXSheetWriter(self.native_book,
                                   self.native_book.active, name)
        else:
            return XLSXSheetWriter(self.native_book,
                                   self.native_book.create_sheet(), name)

    def close(self):
        """
        This call actually save the file
        """
        self.native_book.save(filename=self.file_alike_object)


_xlsx_registry = {
    "file_type": "xlsx",
    "reader": XLSXBook,
    "writer": XLSXWriter,
    "stream_type": "binary",
    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "library": "openpyxl"
}

_xlsm_registry = {
    "file_type": "xlsm",
    "reader": XLSXBook,
    "writer": XLSXWriter,
    "stream_type": "binary",
    "mime_type": "application/vnd.ms-excel.sheet.macroenabled.12",
    "library": "openpyxl"
}

exports = (_xlsx_registry, _xlsm_registry)